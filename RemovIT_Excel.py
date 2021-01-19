import openpyxl
import os
import sys


def error_note():
    workbook_after_del['Arkusz1']['C4'].value = \
        'Bazę danych powinieneś nazwać \"db.xlsx\", a bazę do usunięcia \"del.xlsx\".' \
        ' Obie bazy powinny mieć rekordy w kolumnie A w arkuszu \"Arkusz1\"'
    workbook_after_del.save('db_del.xlsx')
    sys.exit()


os.chdir('./')

workbook_after_del = openpyxl.Workbook()
workbook_after_del.active.title = 'Arkusz1'
try:
    workbook_db = openpyxl.load_workbook('db.xlsx')
except FileNotFoundError:
    error_note()

# noinspection PyUnboundLocalVariable
if workbook_db.sheetnames[0] == "Arkusz1" or workbook_db.sheetnames[0] == 'Sheet1':
    sheet_db = workbook_db[workbook_db.sheetnames[0]]
else:
    error_note()
try:
    workbook_del = openpyxl.load_workbook("del.xlsx")
except FileNotFoundError:
    error_note()

# noinspection PyUnboundLocalVariable
if workbook_del.sheetnames[0] == "Arkusz1" or workbook_del.sheetnames[0] == 'Sheet1':
    sheet_del = workbook_del[workbook_db.sheetnames[0]]
else:
    error_note()

# noinspection PyUnboundLocalVariable
set_db = {i[0] for i in sheet_db.values}
# noinspection PyUnboundLocalVariable
set_del = {i[0] for i in sheet_del.values}
set_result = set_db - set_del

for index, row in enumerate(sorted(set_result)):
    workbook_after_del['Arkusz1'][f"A{index + 1}"].value = row

workbook_after_del['Arkusz1']['D4'].value = 'Liczba rekordów w bazie (bez powtórzeń):'
workbook_after_del['Arkusz1']['E4'].value = f'{len(set_db)}'

workbook_after_del['Arkusz1']['D5'].value = 'Liczba rekordów do usunięcia (bez powtórzeń):'
workbook_after_del['Arkusz1']['E5'].value = f'{len(set_del)}'

workbook_after_del['Arkusz1']['D6'].value = 'Liczba rekordów po usunięciu:'
workbook_after_del['Arkusz1']['E6'].value = f'{len(set_result)}'

workbook_after_del.save('db_del.xlsx')
